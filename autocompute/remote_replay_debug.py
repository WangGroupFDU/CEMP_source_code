
from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
import fnmatch
import importlib
import json
import os
import posixpath
import shlex
import shutil
import subprocess
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import close_old_connections

from autocompute.capability_registry import (
    TASK_TYPE_TO_CAPABILITY,
    get_required_capability,
)
from autocompute.models import ComputeTask
from autocompute.remote_utils import (
    _build_ssh_command,
    _build_remote_target_path,
    _check_remote_server_runtime_health,
    _execute_remote_task_with_result_handling,
    _get_remote_server_by_name,
    _load_remote_server_pool,
    _pull_remote_to_local,
    _ssh_run_batch,
)
from autocompute.failure_utils import normalize_failure_message



REMOTE_QC_DOWNLOADS_TARGET = posixpath.join("AutoCompute", "QcCompute", "Downloads")
REMOTE_MD_DOWNLOADS_TARGET = posixpath.join("AutoCompute", "MDCompute", "Downloads")
REMOTE_MARKOV_DOWNLOADS_TARGET = posixpath.join("AutoCompute", "MarkovAnalysis", "Downloads")
REMOTE_POLYMER_GENERATE_TARGET = posixpath.join("Polymer", "GeneratePolymer")


def get_default_fixture_registry_path() -> str:

    return os.path.join(
        settings.BASE_DIR,
        "static",
        "remote_server_info",
        "remote_task_fixture_registry.json",
    )


def get_default_report_root() -> str:

    return os.path.join(
        settings.MEDIA_ROOT,
        "AutoCompute",
        "NodeReplayDebug",
    )


@dataclass(frozen=True)
class RemoteReplaySpec:

    task_type: str
    func_path: str
    remote_target_subpath: str
    notebooks_to_run: Tuple[str, ...]
    source_mode: str = "fixture_dir"
    copy_fixture_into_download_dir: bool = False
    reset_patterns: Tuple[str, ...] = field(default_factory=tuple)




REPLAY_ALLOWED_FILE_EXTENSIONS: Tuple[str, ...] = (
    ".ipynb",
    ".py",
    ".xls",
    ".xlsx",
    ".xlsm",
)



COMMON_SIGNAL_RESET: Tuple[str, ...] = (
    "success.txt",
    "failure.txt",
)
COMMON_QC_RESET: Tuple[str, ...] = COMMON_SIGNAL_RESET
COMMON_MD_RESET: Tuple[str, ...] = COMMON_SIGNAL_RESET


COMMON_MARKOV_RESET: Tuple[str, ...] = COMMON_SIGNAL_RESET + (
    "markov_controller_result.json",
    "markov_controller_job.json",
    "markov_dispatch.log",
)



MD_REPLAY_SIMULATION_TIME_NS = 5


REMOTE_REPLAY_SPECS: Dict[str, RemoteReplaySpec] = {
    "HTQC_single_point_energy": RemoteReplaySpec(
        task_type="HTQC_single_point_energy",
        func_path="autocompute.remote_utils.run_Gaussian_single_point_energy_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "Gas_component_1_generate_Gaussian_inputfile.ipynb",
            "Gas_component_2_opt+freq_calculation.ipynb",
            "Gas_component_3_opt+freq_failure_correction.ipynb",
            "Gas_component_4_opt+freq_imaginary_frequencies.ipynb",
            "Gas_component_5_energy_calculation.ipynb",
            "Gas_component_6_energy_failure_correction.ipynb",
            "Gas_component_7_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "HTQC_binding_energy": RemoteReplaySpec(
        task_type="HTQC_binding_energy",
        func_path="autocompute.remote_utils.run_Gaussian_binding_energy_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "Gas_component_1_generate_Gaussian_inputfile.ipynb",
            "Gas_component_2_opt+freq_calculation.ipynb",
            "Gas_component_3_opt+freq_failure_correction.ipynb",
            "Gas_component_4_opt+freq_imaginary_frequencies.ipynb",
            "Gas_component_5_energy_calculation.ipynb",
            "Gas_component_6_energy_failure_correction.ipynb",
            "Gas_component_7_Extracting_energy_and_free_energy_corrections.ipynb",
            "Gas_dimer_1_generate_Gaussian_inputfile.ipynb",
            "Gas_dimer_2_opt+freq_calculation.ipynb",
            "Gas_dimer_3_opt+freq_failure_correction.ipynb",
            "Gas_dimer_4_opt+freq_imaginary_frequencies.ipynb",
            "Gas_dimer_5_energy_calculation.ipynb",
            "Gas_dimer_6_energy_failure_correction.ipynb",
            "Gas_dimer_7_Extracting_energy_and_free_energy_corrections.ipynb",
            "Data_processing .ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "HTQC_pka_pkb_calculation": RemoteReplaySpec(
        task_type="HTQC_pka_pkb_calculation",
        func_path="autocompute.remote_utils.run_Gaussian_pka_pkb_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "pkb_DFT_1_generate_Gaussian_inputfile.ipynb",
            "pkb_DFT_2_opt+freq_calculation.ipynb",
            "pkb_DFT_3_opt+freq_failure_correction.ipynb",
            "pkb_DFT_4_opt+freq_imaginary_frequencies.ipynb",
            "pkb_DFT_5_energy_calculation.ipynb",
            "pkb_DFT_6_energy_failure_correction.ipynb",
            "pkb_DFT_7_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "HTQC_ox_red_calculation": RemoteReplaySpec(
        task_type="HTQC_ox_red_calculation",
        func_path="autocompute.remote_utils.run_Gaussian_ox_red_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "ox_red_1_generate_Gaussian_inputfile.ipynb",
            "ox_red_2_opt+freq_calculation.ipynb",
            "ox_red_3_opt+freq_failure_correction.ipynb",
            "ox_red_4_opt+freq_imaginary_frequencies.ipynb",
            "ox_red_5_energy_calculation.ipynb",
            "ox_red_6_energy_failure_correction.ipynb",
            "ox_red_7_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "HTQC_reaction_thermo_properties_calculation": RemoteReplaySpec(
        task_type="HTQC_reaction_thermo_properties_calculation",
        func_path="autocompute.remote_utils.run_Gaussian_reaction_thermo_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "reaction_thermo_1_generate_Gaussian_inputfile.ipynb",
            "reaction_thermo_2_opt+freq_calculation.ipynb",
            "reaction_thermo_3_opt+freq_failure_correction.ipynb",
            "reaction_thermo_4_opt+freq_imaginary_frequencies.ipynb",
            "reaction_thermo_5_energy_calculation.ipynb",
            "reaction_thermo_6_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "HTQC_global_reaction_properties_descriptors_calculation": RemoteReplaySpec(
        task_type="HTQC_global_reaction_properties_descriptors_calculation",
        func_path="autocompute.remote_utils.run_Gaussian_global_reaction_properties_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "reaction_1_generate_Gaussian_inputfile.ipynb",
            "reaction_2_opt+freq_calculation.ipynb",
            "reaction_3_opt+freq_failure_correction.ipynb",
            "reaction_4_opt+freq_imaginary_frequencies.ipynb",
            "reaction_5_energy_calculation.ipynb",
            "reaction_6_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "HTQC_single_point_energy_orca": RemoteReplaySpec(
        task_type="HTQC_single_point_energy_orca",
        func_path="autocompute.remote_utils.run_ORCA_single_point_energy_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "ORCA_Gas_1_generate_ORCA_inputfile.ipynb",
            "ORCA_Gas_2_opt+freq_calculation.ipynb",
            "ORCA_Gas_3_opt+freq_imaginary_frequencies.ipynb",
            "ORCA_Gas_4_energy_calculation.ipynb",
            "ORCA_Gas_5_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "HTQC_binding_energy_orca": RemoteReplaySpec(
        task_type="HTQC_binding_energy_orca",
        func_path="autocompute.remote_utils.run_ORCA_binding_energy_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "ORCA_Gas_component_1_generate_ORCA_inputfile.ipynb",
            "ORCA_Gas_component_2_opt+freq_calculation.ipynb",
            "ORCA_Gas_component_3_opt+freq_imaginary_frequencies.ipynb",
            "ORCA_Gas_component_4_energy_calculation.ipynb",
            "ORCA_Gas_component_5_Extracting_energy_and_free_energy_corrections.ipynb",
            "ORCA_Gas_dimer_1_generate_ORCA_inputfile.ipynb",
            "ORCA_Gas_dimer_2_opt+freq_calculation.ipynb",
            "ORCA_Gas_dimer_3_opt+freq_imaginary_frequencies.ipynb",
            "ORCA_Gas_dimer_4_energy_calculation.ipynb",
            "ORCA_Gas_dimer_5_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "HTQC_ox_red_calculation_orca": RemoteReplaySpec(
        task_type="HTQC_ox_red_calculation_orca",
        func_path="autocompute.remote_utils.run_ORCA_ox_red_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "ORCA_ox_red_1_generate_ORCA_inputfile.ipynb",
            "ORCA_ox_red_2_opt+freq_calculation.ipynb",
            "ORCA_ox_red_3_opt+freq_imaginary_frequencies.ipynb",
            "ORCA_ox_red_4_energy_calculation.ipynb",
            "ORCA_ox_red_5_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "Manual_Mode_QCcompute": RemoteReplaySpec(
        task_type="Manual_Mode_QCcompute",
        func_path="autocompute.remote_utils.run_ORCA_manual_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "ORCA_Gas_2_opt+freq_calculation.ipynb",
            "ORCA_Gas_3_opt+freq_imaginary_frequencies.ipynb",
            "ORCA_Gas_4_energy_calculation.ipynb",
            "ORCA_Gas_5_Extracting_energy_and_free_energy_corrections.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "Manual_Mode_QCcompute_energy": RemoteReplaySpec(
        task_type="Manual_Mode_QCcompute_energy",
        func_path="autocompute.remote_utils.run_ORCA_manual_notebook_tasks_remote_energy",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "ORCA_Gas_1_energy_calculation.ipynb",
            "ORCA_Gas_2_Extracting_energy.ipynb",
        ),
        reset_patterns=COMMON_QC_RESET,
    ),
    "MDCoumpute": RemoteReplaySpec(
        task_type="MDCoumpute",
        func_path="autocompute.remote_utils.run_Gromacs_MD_notebook_tasks_remote",
        remote_target_subpath=REMOTE_MD_DOWNLOADS_TARGET,
        notebooks_to_run=(
            "1_Polymer_RESP_repeat_unit.ipynb",
            "2_Polymer_chg_and_Polymer_creation_ Linear_polymer.ipynb",
            "3_create_Polymer_itp_top.ipynb",
            "4_generate_Gaussian_inputfile.ipynb",
            "5_opt+freq_calculation.ipynb",
            "6_opt+freq_failure_correction.ipynb",
            "7_opt+freq_imaginary_frequencies.ipynb",
            "8_MD_process.ipynb",
            "9_post_analysis.ipynb",
            "11_component_energy_calculation.ipynb",
            "12_calculate_solvent_cage_escape_energy.ipynb",
            "13_coordination_environment_distribution.ipynb",
        ),
        reset_patterns=COMMON_MD_RESET,
    ),
    "DrawESP": RemoteReplaySpec(
        task_type="DrawESP",
        func_path="autocompute.remote_utils.run_draw_ESP_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=("auto_draw_ESP.ipynb",),
        reset_patterns=COMMON_QC_RESET,
    ),
    "DrawESP_remote": RemoteReplaySpec(
        task_type="DrawESP_remote",
        func_path="autocompute.remote_utils.run_draw_ESP_notebook_tasks_gbw_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=("auto_draw_ESP_gbw.ipynb",),
        reset_patterns=COMMON_QC_RESET,
    ),
    "Draw_HOMO_LUMO_orb": RemoteReplaySpec(
        task_type="Draw_HOMO_LUMO_orb",
        func_path="autocompute.remote_utils.run_draw_HOMO_LUMO_orb_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=("draw_HOMO_LUMO_orb.ipynb",),
        reset_patterns=COMMON_QC_RESET,
    ),
    "NCI_analysis": RemoteReplaySpec(
        task_type="NCI_analysis",
        func_path="autocompute.remote_utils.run_NCI_SCF_analysis_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=("NCI_analysis.ipynb",),
        reset_patterns=COMMON_QC_RESET,
    ),
    "NCI_promolecular_analysis": RemoteReplaySpec(
        task_type="NCI_promolecular_analysis",
        func_path="autocompute.remote_utils.run_NCI_promolecular_analysis_notebook_tasks_remote",
        remote_target_subpath=REMOTE_QC_DOWNLOADS_TARGET,
        notebooks_to_run=("NCI_analysis_promolecular.ipynb",),
        reset_patterns=COMMON_QC_RESET,
    ),
    "Markov_GDyNet_analysis": RemoteReplaySpec(
        task_type="Markov_GDyNet_analysis",
        func_path="autocompute.remote_utils.run_markov_gdynet_analysis_remote",
        remote_target_subpath=REMOTE_MARKOV_DOWNLOADS_TARGET,
        notebooks_to_run=(),
        source_mode="markov_gdynet_package",
        copy_fixture_into_download_dir=True,
        reset_patterns=COMMON_MARKOV_RESET,
    ),
    "Generate_homopolymer": RemoteReplaySpec(
        task_type="Generate_homopolymer",
        func_path="polymer.remote_utils.generate_polymer_run_notebook_tasks_remote",
        remote_target_subpath=REMOTE_POLYMER_GENERATE_TARGET,
        notebooks_to_run=(
            "1_Polymer_RESP_repeat_unit.ipynb",
            "2_1_opt+freq_failure_correction.ipynb",
            "2_2_opt+freq_imaginary_frequencies.ipynb",
            "2_Polymer_chg_and_Polymer_creation_Linear_polymer.ipynb",
            "3_create_Polymer_itp_top.ipynb",
        ),
        reset_patterns=COMMON_MD_RESET,
    ),
    "Generate_random_copolymer": RemoteReplaySpec(
        task_type="Generate_random_copolymer",
        func_path="polymer.remote_utils.generate_polymer_run_notebook_tasks_remote",
        remote_target_subpath=REMOTE_POLYMER_GENERATE_TARGET,
        notebooks_to_run=(
            "1_Polymer_RESP_repeat_unit.ipynb",
            "2_1_opt+freq_failure_correction.ipynb",
            "2_2_opt+freq_imaginary_frequencies.ipynb",
            "2_Polymer_chg_and_Polymer_creation_Linear_polymer.ipynb",
            "3_create_Polymer_itp_top.ipynb",
        ),
        reset_patterns=COMMON_MD_RESET,
    ),
    "Generate_block_copolymer": RemoteReplaySpec(
        task_type="Generate_block_copolymer",
        func_path="polymer.remote_utils.generate_polymer_run_notebook_tasks_remote",
        remote_target_subpath=REMOTE_POLYMER_GENERATE_TARGET,
        notebooks_to_run=(
            "1_Polymer_RESP_repeat_unit.ipynb",
            "2_1_opt+freq_failure_correction.ipynb",
            "2_2_opt+freq_imaginary_frequencies.ipynb",
            "2_Polymer_chg_and_Polymer_creation_Linear_polymer.ipynb",
            "3_create_Polymer_itp_top.ipynb",
        ),
        reset_patterns=COMMON_MD_RESET,
    ),
}


@dataclass
class FixtureRecord:

    task_type: str
    capability: str
    source_kind: str
    source_server_name: str
    source_task_id: str
    source_folder_path: str
    source_remote_dir: str
    task_dir_name: str
    notebooks_to_run: List[str]
    func_path: str
    remote_target_subpath: str


@dataclass
class ReplayTaskResult:

    task_type: str
    target_server_name: str
    passed: bool
    preflight_ok: bool
    preflight_error: str
    task_id: str
    local_work_dir: str
    target_remote_dir: str
    success_txt_exists: bool
    failure_txt_exists: bool
    failure_excerpt: str
    started_at: str
    finished_at: str
    elapsed_seconds: float


def get_registered_remote_task_types() -> List[str]:

    return sorted(TASK_TYPE_TO_CAPABILITY.keys())


def validate_replay_spec_coverage() -> Tuple[List[str], List[str]]:

    registered = set(get_registered_remote_task_types())
    specified = set(REMOTE_REPLAY_SPECS.keys())
    missing = sorted(registered - specified)
    stale = sorted(specified - registered)
    return missing, stale


def load_fixture_registry(registry_path: Optional[str] = None) -> Dict[str, Dict[str, Any]]:

    path = registry_path or get_default_fixture_registry_path()
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError("Fixture registry must be a JSON object keyed by task_type.")
    return payload


def save_fixture_registry(records: Dict[str, Dict[str, Any]], registry_path: Optional[str] = None) -> str:

    path = registry_path or get_default_fixture_registry_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(records, handle, ensure_ascii=False, indent=2)
    return path


def _build_source_remote_dir(task: ComputeTask, source_server: Dict[str, Any], spec: RemoteReplaySpec) -> str:

    source_base_dir = _build_remote_target_path(source_server, spec.remote_target_subpath)
    task_dir_name = os.path.basename(str(task.folder_path).rstrip("/"))
    return posixpath.join(source_base_dir.rstrip("/"), task_dir_name)


def _remote_success_marker_exists(source_server: Dict[str, Any], remote_dir: str) -> bool:

    try:
        _ssh_run_batch(
            source_server["IP"],
            f"test -f {json.dumps(posixpath.join(remote_dir, 'success.txt'))}",
            timeout=30,
        )
    except Exception:
        return False
    return True


def _local_success_marker_exists(folder_path: str) -> bool:

    success_path = os.path.join(str(folder_path), "success.txt")
    return os.path.isfile(success_path)


def _find_remote_fixture_across_server_pool(
    *,
    task_dir_name: str,
    spec: RemoteReplaySpec,
) -> Optional[Tuple[Dict[str, Any], str]]:

    for server in _load_remote_server_pool():
        candidate_dir = posixpath.join(
            _build_remote_target_path(server, spec.remote_target_subpath).rstrip("/"),
            task_dir_name,
        )
        if _remote_success_marker_exists(server, candidate_dir):
            return server, candidate_dir
    return None


def discover_fixture_records(
    *,
    registry_path: Optional[str] = None,
) -> Tuple[Dict[str, Dict[str, Any]], List[str]]:

    missing_specs, stale_specs = validate_replay_spec_coverage()
    if missing_specs or stale_specs:
        raise ValueError(
            f"Replay spec coverage mismatch. missing={missing_specs}, stale={stale_specs}"
        )

    records: Dict[str, Dict[str, Any]] = {}
    missing_task_types: List[str] = []
    for task_type in get_registered_remote_task_types():
        spec = REMOTE_REPLAY_SPECS[task_type]
        capability = get_required_capability(task_type)
        selected_record: Optional[FixtureRecord] = None
        candidates = (
            ComputeTask.objects.filter(
                task_type=task_type,
                status="success",
                remote_type="remote",
            )
            .exclude(server_name__isnull=True)
            .exclude(server_name__exact="")
            .order_by("-created_at")
        )
        for task in candidates:
            source_server = _get_remote_server_by_name(str(task.server_name))
            if source_server is None:
                continue
            source_remote_dir = _build_source_remote_dir(task, source_server, spec)
            if not _remote_success_marker_exists(source_server, source_remote_dir):
                continue
            selected_record = FixtureRecord(
                task_type=task_type,
                capability=capability or "",
                source_kind="remote_server",
                source_server_name=str(task.server_name),
                source_task_id=str(task.task_id),
                source_folder_path=str(task.folder_path),
                source_remote_dir=source_remote_dir,
                task_dir_name=os.path.basename(str(task.folder_path).rstrip("/")),
                notebooks_to_run=list(spec.notebooks_to_run),
                func_path=spec.func_path,
                remote_target_subpath=spec.remote_target_subpath,
            )
            break

        if selected_record is None:
            remote_candidates_without_bound_server = (
                ComputeTask.objects.filter(
                    task_type=task_type,
                    status="success",
                    remote_type="remote",
                )
                .exclude(folder_path__isnull=True)
                .exclude(folder_path__exact="")
                .order_by("-created_at")
            )
            for task in remote_candidates_without_bound_server:
                task_dir_name = os.path.basename(str(task.folder_path).rstrip("/"))
                located = _find_remote_fixture_across_server_pool(
                    task_dir_name=task_dir_name,
                    spec=spec,
                )
                if located is None:
                    continue
                located_server, source_remote_dir = located
                selected_record = FixtureRecord(
                    task_type=task_type,
                    capability=capability or "",
                    source_kind="remote_server",
                    source_server_name=str(located_server["server_name"]),
                    source_task_id=str(task.task_id),
                    source_folder_path=str(task.folder_path),
                    source_remote_dir=source_remote_dir,
                    task_dir_name=task_dir_name,
                    notebooks_to_run=list(spec.notebooks_to_run),
                    func_path=spec.func_path,
                    remote_target_subpath=spec.remote_target_subpath,
                )
                break

        if selected_record is None:
            local_candidates = (
                ComputeTask.objects.filter(
                    task_type=task_type,
                    status="success",
                )
                .exclude(folder_path__isnull=True)
                .exclude(folder_path__exact="")
                .order_by("-created_at")
            )
            for task in local_candidates:
                local_dir = str(task.folder_path)
                if not _local_success_marker_exists(local_dir):
                    continue
                selected_record = FixtureRecord(
                    task_type=task_type,
                    capability=capability or "",
                    source_kind="local_fs",
                    source_server_name=str(task.server_name or ""),
                    source_task_id=str(task.task_id),
                    source_folder_path=local_dir,
                    source_remote_dir="",
                    task_dir_name=os.path.basename(local_dir.rstrip("/")),
                    notebooks_to_run=list(spec.notebooks_to_run),
                    func_path=spec.func_path,
                    remote_target_subpath=spec.remote_target_subpath,
                )
                break

        if selected_record is None:
            missing_task_types.append(task_type)
            continue

        records[task_type] = asdict(selected_record)

    save_fixture_registry(records, registry_path=registry_path)
    return records, missing_task_types


def _import_callable(func_path: str):

    module_name, func_name = func_path.rsplit(".", 1)
    module = importlib.import_module(module_name)
    return getattr(module, func_name)


def _ensure_empty_dir(path: str) -> None:

    shutil.rmtree(path, ignore_errors=True)
    os.makedirs(path, exist_ok=True)


def _copy_tree_contents(src: str, dst: str) -> None:

    os.makedirs(dst, exist_ok=True)
    for entry in os.listdir(src):
        src_path = os.path.join(src, entry)
        dst_path = os.path.join(dst, entry)
        if os.path.isdir(src_path):
            shutil.copytree(src_path, dst_path)
        else:
            shutil.copy2(src_path, dst_path)


def _should_reset_path(root_dir: str, path: str, patterns: Sequence[str]) -> bool:

    rel_path = os.path.relpath(path, root_dir).replace("\\", "/")
    base_name = os.path.basename(path)
    for pattern in patterns:
        normalized_pattern = pattern.replace("\\", "/")
        if fnmatch.fnmatch(rel_path, normalized_pattern) or fnmatch.fnmatch(base_name, normalized_pattern):
            return True
    return False


def _is_replay_allowed_file(path: str) -> bool:

    lower_name = os.path.basename(path).lower()
    return lower_name.endswith(REPLAY_ALLOWED_FILE_EXTENSIONS)


def reset_fixture_tree(root_dir: str, spec: RemoteReplaySpec) -> List[str]:

    removed_paths: List[str] = []
    if not os.path.isdir(root_dir):
        return removed_paths

    file_paths: List[str] = []
    dir_paths: List[str] = []
    for current_root, dirnames, filenames in os.walk(root_dir):
        for filename in filenames:
            file_paths.append(os.path.join(current_root, filename))
        for dirname in dirnames:
            dir_paths.append(os.path.join(current_root, dirname))

    
    for path in sorted(file_paths, key=lambda item: item.count(os.sep), reverse=True):
        if _is_replay_allowed_file(path):
            continue
        rel_path = os.path.relpath(path, root_dir).replace("\\", "/")
        try:
            os.remove(path)
        except FileNotFoundError:
            continue
        removed_paths.append(rel_path)

    
    
    for path in sorted(dir_paths, key=lambda item: item.count(os.sep), reverse=True):
        if not os.path.isdir(path):
            continue
        try:
            if not os.listdir(path):
                rel_path = os.path.relpath(path, root_dir).replace("\\", "/")
                os.rmdir(path)
                removed_paths.append(rel_path)
        except FileNotFoundError:
            continue
    return removed_paths


def _rewrite_system_xlsx_time_ns(system_xlsx_path: str, target_time_ns: int) -> Dict[str, Any]:

    from openpyxl import load_workbook

    if not os.path.exists(system_xlsx_path):
        raise FileNotFoundError(f"System.xlsx not found: {system_xlsx_path}")

    workbook = load_workbook(system_xlsx_path)
    worksheet = workbook.active

    header_to_column: Dict[str, int] = {}
    for cell in worksheet[1]:
        if cell.value is None:
            continue
        header_to_column[str(cell.value).strip()] = int(cell.column)

    target_column = header_to_column.get("time (ns)")
    if target_column is None:
        raise ValueError(f"'time (ns)' column not found in {system_xlsx_path}")

    updated_rows = 0
    for row_index in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_index, column=target_column)
        if cell.value is None or str(cell.value).strip() == "":
            continue
        cell.value = target_time_ns
        updated_rows += 1

    if updated_rows == 0:
        raise ValueError(f"No non-empty 'time (ns)' cells found in {system_xlsx_path}")

    workbook.save(system_xlsx_path)
    return {
        "system_xlsx_path": system_xlsx_path,
        "column_name": "time (ns)",
        "target_time_ns": target_time_ns,
        "updated_rows": updated_rows,
    }


def apply_fixture_replay_mutations(root_dir: str, spec: RemoteReplaySpec) -> Dict[str, Any]:

    if spec.task_type != "MDCoumpute":
        return {}

    system_xlsx_path = os.path.join(root_dir, "System.xlsx")
    return {
        "md_time_override": _rewrite_system_xlsx_time_ns(
            system_xlsx_path=system_xlsx_path,
            target_time_ns=MD_REPLAY_SIMULATION_TIME_NS,
        )
    }


def _get_debug_user():

    user_model = get_user_model()
    user, _ = user_model.objects.get_or_create(
        username="node_debug_replay",
        defaults={"email": "user@example.com"},
    )
    if not user.has_usable_password():
        return user
    user.set_unusable_password()
    user.save(update_fields=["password"])
    return user


def _write_failure_file_if_missing(download_dir: str, message: str) -> None:

    failure_path = os.path.join(download_dir, "failure.txt")
    if os.path.exists(failure_path):
        return
    os.makedirs(download_dir, exist_ok=True)
    with open(failure_path, "w", encoding="utf-8") as handle:
        handle.write(normalize_failure_message(message).rstrip() + "\n")


def _collect_failure_excerpt(download_dir: str, limit: int = 2000) -> str:

    failure_path = os.path.join(download_dir, "failure.txt")
    if not os.path.exists(failure_path):
        return ""
    with open(failure_path, "r", encoding="utf-8", errors="ignore") as handle:
        return handle.read(limit)


def _build_replay_remote_target_root(
    *,
    target_server: Dict[str, Any],
    spec: RemoteReplaySpec,
    task_type: str,
    target_remote_root: Optional[str],
) -> str:

    custom_root = str(target_remote_root or "").strip().replace("\\", "/").rstrip("/")
    if custom_root:
        return posixpath.join(custom_root, task_type)
    return _build_remote_target_path(target_server, spec.remote_target_subpath)


def _resolve_replay_source_dir(spec: RemoteReplaySpec, cleaned_fixture_dir: str) -> str:

    if spec.source_mode == "fixture_dir":
        return cleaned_fixture_dir
    if spec.source_mode == "markov_gdynet_package":
        gdynet_dir = str(getattr(settings, "MARKOV_GDYNET_PACKAGE_DIR", "") or "").strip()
        if not gdynet_dir:
            raise ValueError("MARKOV_GDYNET_PACKAGE_DIR is not configured.")
        return gdynet_dir
    raise ValueError(f"Unsupported replay source_mode: {spec.source_mode}")


def _build_task_id(task_type: str, target_server_name: str) -> str:

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    random_suffix = uuid.uuid4().hex[:10]
    return f"node_debug__{task_type}__{target_server_name}__{timestamp}__{random_suffix}"


def _build_report_markdown(
    *,
    target_server_name: str,
    records: Sequence[ReplayTaskResult],
    missing_task_types: Sequence[str],
) -> str:

    passed = [item for item in records if item.passed]
    failed = [item for item in records if not item.passed]
    lines: List[str] = [
        f"# Remote Replay Qualification Report - {target_server_name}",
        "",
        f"- Generated At: {datetime.now().isoformat()}",
        f"- Total Tested: {len(records)}",
        f"- Passed: {len(passed)}",
        f"- Failed: {len(failed)}",
    ]
    if missing_task_types:
        lines.extend(
            [
                "",
                "## Missing Fixtures",
                "",
                *[f"- {task_type}" for task_type in missing_task_types],
            ]
        )
    lines.extend(["", "## Results", ""])
    for item in records:
        lines.extend(
            [
                f"### {item.task_type}",
                "",
                f"- Passed: `{item.passed}`",
                f"- Preflight OK: `{item.preflight_ok}`",
                f"- Task ID: `{item.task_id}`",
                f"- Local Work Dir: `{item.local_work_dir}`",
                f"- Target Remote Dir: `{item.target_remote_dir}`",
                f"- success.txt: `{item.success_txt_exists}`",
                f"- failure.txt: `{item.failure_txt_exists}`",
                f"- Elapsed Seconds: `{item.elapsed_seconds:.2f}`",
            ]
        )
        if item.preflight_error:
            lines.append(f"- Preflight Error: `{item.preflight_error}`")
        if item.failure_excerpt:
            lines.extend(
                [
                    "",
                    "```text",
                    item.failure_excerpt.rstrip(),
                    "```",
                ]
            )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _ssh_run_noninteractive_for_replay(
    remote_login: str,
    cmd: str,
    timeout=None,
    *,
    ssh_port: Optional[int] = None,
    server_info_file_path: Optional[str] = None,
):

    
    
    
    
    
    script = "unset LC_ALL LC_CTYPE LANGUAGE || true\n"
    script += "export LANG=en_US.UTF-8\n"
    script += "set -euo pipefail\n"
    script += cmd.rstrip() + "\n"

    ssh_cmd = _build_ssh_command(
        remote_login,
        batch_mode=True,
        allocate_tty=False,
        ssh_port=ssh_port,
        server_info_file_path=server_info_file_path,
    )
    remote_command = f"bash --noprofile --norc -lc {shlex.quote(script)}"
    
    
    local_env = os.environ.copy()
    local_env.pop("LC_ALL", None)
    local_env.pop("LC_CTYPE", None)
    local_env.pop("LANGUAGE", None)
    local_env["LANG"] = "en_US.UTF-8"
    return subprocess.run(
        ssh_cmd + [remote_command],
        text=True,
        capture_output=True,
        check=True,
        timeout=timeout,
        env=local_env,
    )


@contextmanager
def _patched_replay_remote_execution():

    import autocompute.remote_utils as autocompute_remote_utils_module
    import polymer.remote_utils as polymer_remote_utils_module

    original_autocompute_ssh_run = autocompute_remote_utils_module._ssh_run
    original_polymer_ssh_run = polymer_remote_utils_module._ssh_run

    autocompute_remote_utils_module._ssh_run = _ssh_run_noninteractive_for_replay
    polymer_remote_utils_module._ssh_run = _ssh_run_noninteractive_for_replay
    try:
        yield
    finally:
        autocompute_remote_utils_module._ssh_run = original_autocompute_ssh_run
        polymer_remote_utils_module._ssh_run = original_polymer_ssh_run


def _run_single_fixture_replay(
    *,
    task_type: str,
    fixture: Dict[str, Any],
    spec: RemoteReplaySpec,
    target_server_name: str,
    target_server: Dict[str, Any],
    stage_root: str,
    debug_user,
    target_remote_root: Optional[str] = None,
) -> ReplayTaskResult:

    close_old_connections()
    started_at_dt = datetime.now()
    started_at = started_at_dt.isoformat()

    source_snapshot_dir = os.path.join(stage_root, "source_snapshot")
    cleaned_source_dir = os.path.join(stage_root, "source_clean")
    local_download_dir = os.path.join(stage_root, "download")

    _ensure_empty_dir(stage_root)
    _ensure_empty_dir(source_snapshot_dir)
    _ensure_empty_dir(cleaned_source_dir)
    _ensure_empty_dir(local_download_dir)

    task_id = _build_task_id(task_type, target_server_name)
    actual_remote_target = _build_replay_remote_target_root(
        target_server=target_server,
        spec=spec,
        task_type=task_type,
        target_remote_root=target_remote_root,
    )
    target_remote_dir = posixpath.join(
        actual_remote_target.rstrip("/"),
        os.path.basename(local_download_dir.rstrip("/")),
    )

    required_capability = get_required_capability(task_type)
    server_capabilities = set(target_server.get("capabilities") or [])
    if required_capability and required_capability not in server_capabilities:
        preflight_error = (
            f"Target server {target_server_name} is not registered for capability "
            f"{required_capability} required by {task_type}."
        )
        _write_failure_file_if_missing(local_download_dir, preflight_error)
        close_old_connections()
        return ReplayTaskResult(
            task_type=task_type,
            target_server_name=target_server_name,
            passed=False,
            preflight_ok=False,
            preflight_error=preflight_error,
            task_id=task_id,
            local_work_dir=local_download_dir,
            target_remote_dir=target_remote_dir,
            success_txt_exists=False,
            failure_txt_exists=True,
            failure_excerpt=_collect_failure_excerpt(local_download_dir),
            started_at=started_at,
            finished_at=datetime.now().isoformat(),
            elapsed_seconds=(datetime.now() - started_at_dt).total_seconds(),
        )

    preflight_ok, preflight_error = _check_remote_server_runtime_health(target_server, task_type)
    if not preflight_ok:
        _write_failure_file_if_missing(
            local_download_dir,
            normalize_failure_message(preflight_error or "Runtime health check failed."),
        )
        close_old_connections()
        return ReplayTaskResult(
            task_type=task_type,
            target_server_name=target_server_name,
            passed=False,
            preflight_ok=False,
            preflight_error=preflight_error,
            task_id=task_id,
            local_work_dir=local_download_dir,
            target_remote_dir=target_remote_dir,
            success_txt_exists=False,
            failure_txt_exists=True,
            failure_excerpt=_collect_failure_excerpt(local_download_dir),
            started_at=started_at,
            finished_at=datetime.now().isoformat(),
            elapsed_seconds=(datetime.now() - started_at_dt).total_seconds(),
        )

    source_kind = str(fixture.get("source_kind") or "remote_server").strip()
    source_server = None
    if source_kind == "remote_server":
        source_server = _get_remote_server_by_name(fixture["source_server_name"])
        if source_server is None:
            preflight_error = f"Fixture source server {fixture['source_server_name']} is not configured."
            _write_failure_file_if_missing(local_download_dir, preflight_error)
            close_old_connections()
            return ReplayTaskResult(
                task_type=task_type,
                target_server_name=target_server_name,
                passed=False,
                preflight_ok=True,
                preflight_error=preflight_error,
                task_id=task_id,
                local_work_dir=local_download_dir,
                target_remote_dir=target_remote_dir,
                success_txt_exists=False,
                failure_txt_exists=True,
                failure_excerpt=_collect_failure_excerpt(local_download_dir),
                started_at=started_at,
                finished_at=datetime.now().isoformat(),
                elapsed_seconds=(datetime.now() - started_at_dt).total_seconds(),
            )
    elif source_kind != "local_fs":
        preflight_error = f"Unsupported fixture source_kind: {source_kind}"
        _write_failure_file_if_missing(local_download_dir, preflight_error)
        close_old_connections()
        return ReplayTaskResult(
            task_type=task_type,
            target_server_name=target_server_name,
            passed=False,
            preflight_ok=True,
            preflight_error=preflight_error,
            task_id=task_id,
            local_work_dir=local_download_dir,
            target_remote_dir=target_remote_dir,
            success_txt_exists=False,
            failure_txt_exists=True,
            failure_excerpt=_collect_failure_excerpt(local_download_dir),
            started_at=started_at,
            finished_at=datetime.now().isoformat(),
            elapsed_seconds=(datetime.now() - started_at_dt).total_seconds(),
        )

    task = ComputeTask.objects.create(
        user=debug_user,
        task_type=task_type,
        task_id=task_id,
        folder_path=local_download_dir,
        status="pending",
        remote_type="remote",
        server_name=target_server_name,
        priority=1,
        status_message="Remote replay debug task created by node qualification workflow.",
    )

    replay_exception = ""
    try:
        if source_kind == "remote_server":
            _pull_remote_to_local(
                source_server["IP"],
                fixture["source_remote_dir"],
                source_snapshot_dir,
            )
        else:
            _copy_tree_contents(str(fixture["source_folder_path"]), source_snapshot_dir)
        _copy_tree_contents(source_snapshot_dir, cleaned_source_dir)
        reset_fixture_tree(cleaned_source_dir, spec)
        apply_fixture_replay_mutations(cleaned_source_dir, spec)
        if spec.copy_fixture_into_download_dir:
            _copy_tree_contents(cleaned_source_dir, local_download_dir)

        source_dir_arg = _resolve_replay_source_dir(spec, cleaned_source_dir)
        task_func = _import_callable(spec.func_path)
        _execute_remote_task_with_result_handling(
            task_func=task_func,
            source_dir=source_dir_arg,
            download_dir=local_download_dir,
            task=task,
            remote_target=actual_remote_target,
            remote_login=target_server["IP"],
        )
    except Exception as exc:  
        replay_exception = f"{type(exc).__name__}: {exc}"
        _write_failure_file_if_missing(
            local_download_dir,
            normalize_failure_message(replay_exception),
        )
        task.refresh_from_db()
        task.status = "failed"
        task.status_message = replay_exception
        task.save(update_fields=["status", "status_message"])
    finally:
        success_txt_exists = os.path.exists(os.path.join(local_download_dir, "success.txt"))
        failure_txt_exists = os.path.exists(os.path.join(local_download_dir, "failure.txt"))
        if not success_txt_exists and not failure_txt_exists:
            _write_failure_file_if_missing(
                local_download_dir,
                normalize_failure_message(
                    replay_exception or "Replay finished without success.txt or failure.txt."
                ),
            )
            task.refresh_from_db()
            task.status = "failed"
            task.status_message = replay_exception or "Replay finished without success.txt."
            task.save(update_fields=["status", "status_message"])

    result = ReplayTaskResult(
        task_type=task_type,
        target_server_name=target_server_name,
        passed=os.path.exists(os.path.join(local_download_dir, "success.txt")),
        preflight_ok=True,
        preflight_error="",
        task_id=task_id,
        local_work_dir=local_download_dir,
        target_remote_dir=target_remote_dir,
        success_txt_exists=os.path.exists(os.path.join(local_download_dir, "success.txt")),
        failure_txt_exists=os.path.exists(os.path.join(local_download_dir, "failure.txt")),
        failure_excerpt=_collect_failure_excerpt(local_download_dir),
        started_at=started_at,
        finished_at=datetime.now().isoformat(),
        elapsed_seconds=(datetime.now() - started_at_dt).total_seconds(),
    )
    close_old_connections()
    return result


def run_fixture_replays(
    *,
    target_server_name: str,
    registry_path: Optional[str] = None,
    task_types: Optional[Sequence[str]] = None,
    report_root: Optional[str] = None,
    target_remote_root: Optional[str] = None,
    parallel_workers: int = 1,
    continue_on_failure: bool = True,
) -> Dict[str, Any]:

    missing_specs, stale_specs = validate_replay_spec_coverage()
    if missing_specs or stale_specs:
        raise ValueError(
            f"Replay spec coverage mismatch. missing={missing_specs}, stale={stale_specs}"
        )

    registry = load_fixture_registry(registry_path=registry_path)
    requested_task_types = list(task_types or get_registered_remote_task_types())
    missing_task_types = [task_type for task_type in requested_task_types if task_type not in registry]

    target_server = _get_remote_server_by_name(target_server_name)
    if target_server is None:
        raise ValueError(f"Target server {target_server_name} is not configured.")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_root_dir = report_root or get_default_report_root()
    session_root = os.path.join(report_root_dir, target_server_name, timestamp)
    os.makedirs(session_root, exist_ok=True)

    results: List[ReplayTaskResult] = []
    debug_user = _get_debug_user()
    executable_task_types = [task_type for task_type in requested_task_types if task_type in registry]
    max_workers = max(1, int(parallel_workers or 1))

    if not continue_on_failure:
        max_workers = 1

    with _patched_replay_remote_execution():
        if max_workers == 1:
            for task_type in executable_task_types:
                result = _run_single_fixture_replay(
                    task_type=task_type,
                    fixture=registry[task_type],
                    spec=REMOTE_REPLAY_SPECS[task_type],
                    target_server_name=target_server_name,
                    target_server=target_server,
                    stage_root=os.path.join(session_root, task_type),
                    debug_user=debug_user,
                    target_remote_root=target_remote_root,
                )
                results.append(result)
                if not result.passed and not continue_on_failure:
                    break
        else:
            indexed_results: Dict[int, ReplayTaskResult] = {}
            with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="cemp_replay") as executor:
                future_map = {
                    executor.submit(
                        _run_single_fixture_replay,
                        task_type=task_type,
                        fixture=registry[task_type],
                        spec=REMOTE_REPLAY_SPECS[task_type],
                        target_server_name=target_server_name,
                        target_server=target_server,
                        stage_root=os.path.join(session_root, task_type),
                        debug_user=debug_user,
                        target_remote_root=target_remote_root,
                    ): index
                    for index, task_type in enumerate(executable_task_types)
                }

                for future in as_completed(future_map):
                    index = future_map[future]
                    indexed_results[index] = future.result()

            results = [indexed_results[index] for index in sorted(indexed_results.keys())]

    summary = {
        "target_server_name": target_server_name,
        "generated_at": datetime.now().isoformat(),
        "requested_task_types": requested_task_types,
        "missing_task_types": missing_task_types,
        "parallel_workers": max_workers,
        "target_remote_root": target_remote_root or "",
        "total": len(results),
        "passed_count": sum(1 for item in results if item.passed),
        "failed_count": sum(1 for item in results if not item.passed),
        "results": [asdict(item) for item in results],
    }

    summary_json_path = os.path.join(session_root, "summary.json")
    with open(summary_json_path, "w", encoding="utf-8") as handle:
        json.dump(summary, handle, ensure_ascii=False, indent=2)

    summary_md_path = os.path.join(session_root, "summary.md")
    with open(summary_md_path, "w", encoding="utf-8") as handle:
        handle.write(
            _build_report_markdown(
                target_server_name=target_server_name,
                records=results,
                missing_task_types=missing_task_types,
            )
        )

    summary["summary_json_path"] = summary_json_path
    summary["summary_md_path"] = summary_md_path
    summary["all_passed"] = (not missing_task_types) and all(item.passed for item in results)
    return summary

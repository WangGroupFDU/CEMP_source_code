import json
import os
import shutil
import tempfile
from datetime import timedelta
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.utils import timezone

from autocompute.models import ComputeTask
import autocompute.remote_replay_debug as replay_module
from autocompute.remote_replay_debug import (
    REMOTE_REPLAY_SPECS,
    discover_fixture_records,
    run_fixture_replays,
    save_fixture_registry,
    validate_replay_spec_coverage,
)


class RemoteReplayDebugTests(TestCase):

    def setUp(self):
        self.temp_dir = tempfile.mkdtemp(prefix="cemp_replay_debug_")
        self.registry_path = os.path.join(self.temp_dir, "fixture_registry.json")
        self.report_root = os.path.join(self.temp_dir, "reports")
        self.media_root = os.path.join(self.temp_dir, "media")
        os.makedirs(self.media_root, exist_ok=True)
        self.user = get_user_model().objects.create_user(
            username="replay_debug_user",
            password="testpass123",
            email="user@example.com",
        )

    def tearDown(self):
        shutil.rmtree(self.temp_dir, ignore_errors=True)

    def test_validate_replay_spec_coverage_is_complete(self):

        missing, stale = validate_replay_spec_coverage()
        self.assertEqual(missing, [])
        self.assertEqual(stale, [])

    def test_discover_fixture_records_selects_success_task_with_remote_marker(self):

        newer_task = ComputeTask.objects.create(
            user=self.user,
            task_type="HTQC_single_point_energy",
            task_id="newer_task",
            folder_path="/tmp/newer_fixture_dir",
            status="success",
            remote_type="remote",
            server_name="server_a",
        )
        older_task = ComputeTask.objects.create(
            user=self.user,
            task_type="HTQC_single_point_energy",
            task_id="older_task",
            folder_path="/tmp/older_fixture_dir",
            status="success",
            remote_type="remote",
            server_name="server_a",
        )
        ComputeTask.objects.filter(pk=older_task.pk).update(
            created_at=timezone.now() - timedelta(hours=2)
        )
        ComputeTask.objects.filter(pk=newer_task.pk).update(
            created_at=timezone.now() - timedelta(minutes=10)
        )
        older_task.refresh_from_db()
        newer_task.refresh_from_db()

        server = {
            "server_name": "server_a",
            "IP": "user@<PRIVATE_IP>:",
            "remote_target_dir": "/srv/cemp",
            "ssh_port": 22,
            "enabled": True,
            "capabilities": ["gaussian_htqc"],
            "remote_login": "user@<PRIVATE_IP>",
            "task_limit": 3,
            "order": 0,
        }

        with mock.patch(
            "autocompute.remote_replay_debug._get_remote_server_by_name",
            return_value=server,
        ), mock.patch(
            "autocompute.remote_replay_debug._remote_success_marker_exists",
            side_effect=[False, True],
        ):
            records, missing = discover_fixture_records(registry_path=self.registry_path)

        self.assertIn("HTQC_single_point_energy", records)
        selected = records["HTQC_single_point_energy"]
        self.assertEqual(selected["source_task_id"], "older_task")
        self.assertEqual(
            selected["source_remote_dir"],
            "/srv/cemp/AutoCompute/QcCompute/Downloads/older_fixture_dir",
        )
        self.assertIn("HTQC_binding_energy", missing)
        stored = json.loads(Path(self.registry_path).read_text(encoding="utf-8"))
        self.assertEqual(stored["HTQC_single_point_energy"]["source_task_id"], "older_task")
        self.assertTrue(newer_task.created_at >= older_task.created_at)

    @override_settings(MEDIA_ROOT="/tmp")
    def test_run_fixture_replays_marks_success_when_wrapper_writes_success_txt(self):

        registry = {
            "HTQC_single_point_energy": {
                "task_type": "HTQC_single_point_energy",
                "capability": "gaussian_htqc",
                "source_server_name": "source_server",
                "source_task_id": "fixture_task_id",
                "source_folder_path": "/tmp/source_fixture",
                "source_remote_dir": "/srv/source/AutoCompute/QcCompute/Downloads/source_fixture",
                "task_dir_name": "source_fixture",
                "notebooks_to_run": list(REMOTE_REPLAY_SPECS["HTQC_single_point_energy"].notebooks_to_run),
                "func_path": REMOTE_REPLAY_SPECS["HTQC_single_point_energy"].func_path,
                "remote_target_subpath": REMOTE_REPLAY_SPECS["HTQC_single_point_energy"].remote_target_subpath,
            }
        }
        save_fixture_registry(registry, registry_path=self.registry_path)

        source_server = {
            "server_name": "source_server",
            "IP": "user@<PRIVATE_IP>:",
            "remote_target_dir": "/srv/source",
            "ssh_port": 22,
            "enabled": True,
            "capabilities": ["gaussian_htqc"],
            "remote_login": "user@<PRIVATE_IP>",
            "task_limit": 3,
            "order": 0,
        }
        target_server = {
            "server_name": "target_server",
            "IP": "user@<PRIVATE_IP>:",
            "remote_target_dir": "/srv/target",
            "ssh_port": 22,
            "enabled": True,
            "capabilities": ["gaussian_htqc"],
            "remote_login": "user@<PRIVATE_IP>",
            "task_limit": 3,
            "order": 1,
        }

        def fake_pull(_remote_ip, _remote_work_dir, local_dir):
            os.makedirs(local_dir, exist_ok=True)
            Path(local_dir, "success.txt").write_text("old success\n", encoding="utf-8")
            Path(local_dir, "Gas_component_1_generate_Gaussian_inputfile.ipynb").write_text("{}", encoding="utf-8")

        def fake_execute(*, task_func, source_dir, download_dir, task, remote_target, remote_login):
            os.makedirs(download_dir, exist_ok=True)
            Path(download_dir, "success.txt").write_text("new success\n", encoding="utf-8")
            task.status = "success"
            task.save(update_fields=["status"])
            return task

        def fake_get_server(server_name, server_info_file_path=None):  
            if server_name == "source_server":
                return source_server
            if server_name == "target_server":
                return target_server
            return None

        with override_settings(MEDIA_ROOT=self.media_root), mock.patch(
            "autocompute.remote_replay_debug._get_remote_server_by_name",
            side_effect=fake_get_server,
        ), mock.patch(
            "autocompute.remote_replay_debug._check_remote_server_runtime_health",
            return_value=(True, ""),
        ), mock.patch(
            "autocompute.remote_replay_debug._pull_remote_to_local",
            side_effect=fake_pull,
        ), mock.patch(
            "autocompute.remote_replay_debug._execute_remote_task_with_result_handling",
            side_effect=fake_execute,
        ):
            summary = run_fixture_replays(
                target_server_name="target_server",
                registry_path=self.registry_path,
                task_types=["HTQC_single_point_energy"],
                report_root=self.report_root,
                target_remote_root="/srv/target/test_folder",
            )

        self.assertTrue(summary["all_passed"])
        self.assertEqual(summary["passed_count"], 1)
        self.assertTrue(os.path.exists(summary["summary_json_path"]))
        self.assertTrue(os.path.exists(summary["summary_md_path"]))
        result = summary["results"][0]
        self.assertTrue(result["success_txt_exists"])
        self.assertFalse(result["failure_txt_exists"])
        self.assertTrue(
            result["target_remote_dir"].startswith(
                "/srv/target/test_folder/HTQC_single_point_energy/"
            )
        )

    def test_run_fixture_replays_backfills_failure_when_wrapper_returns_without_signals(self):

        registry = {
            "HTQC_single_point_energy": {
                "task_type": "HTQC_single_point_energy",
                "capability": "gaussian_htqc",
                "source_server_name": "source_server",
                "source_task_id": "fixture_task_id",
                "source_folder_path": "/tmp/source_fixture",
                "source_remote_dir": "/srv/source/AutoCompute/QcCompute/Downloads/source_fixture",
                "task_dir_name": "source_fixture",
                "notebooks_to_run": list(REMOTE_REPLAY_SPECS["HTQC_single_point_energy"].notebooks_to_run),
                "func_path": REMOTE_REPLAY_SPECS["HTQC_single_point_energy"].func_path,
                "remote_target_subpath": REMOTE_REPLAY_SPECS["HTQC_single_point_energy"].remote_target_subpath,
            }
        }
        save_fixture_registry(registry, registry_path=self.registry_path)

        source_server = {
            "server_name": "source_server",
            "IP": "user@<PRIVATE_IP>:",
            "remote_target_dir": "/srv/source",
            "ssh_port": 22,
            "enabled": True,
            "capabilities": ["gaussian_htqc"],
            "remote_login": "user@<PRIVATE_IP>",
            "task_limit": 3,
            "order": 0,
        }
        target_server = {
            "server_name": "target_server",
            "IP": "user@<PRIVATE_IP>:",
            "remote_target_dir": "/srv/target",
            "ssh_port": 22,
            "enabled": True,
            "capabilities": ["gaussian_htqc"],
            "remote_login": "user@<PRIVATE_IP>",
            "task_limit": 3,
            "order": 1,
        }

        def fake_pull(_remote_ip, _remote_work_dir, local_dir):
            os.makedirs(local_dir, exist_ok=True)
            Path(local_dir, "Gas_component_1_generate_Gaussian_inputfile.ipynb").write_text("{}", encoding="utf-8")


        def fake_execute(*, task_func, source_dir, download_dir, task, remote_target, remote_login):  
            os.makedirs(download_dir, exist_ok=True)
            task.status = "pending"
            task.save(update_fields=["status"])
            return task

        def fake_get_server(server_name, server_info_file_path=None):  
            if server_name == "source_server":
                return source_server
            if server_name == "target_server":
                return target_server
            return None

        with override_settings(MEDIA_ROOT=self.media_root), mock.patch(
            "autocompute.remote_replay_debug._get_remote_server_by_name",
            side_effect=fake_get_server,
        ), mock.patch(
            "autocompute.remote_replay_debug._check_remote_server_runtime_health",
            return_value=(True, ""),
        ), mock.patch(
            "autocompute.remote_replay_debug._pull_remote_to_local",
            side_effect=fake_pull,
        ), mock.patch(
            "autocompute.remote_replay_debug._execute_remote_task_with_result_handling",
            side_effect=fake_execute,
        ):
            summary = run_fixture_replays(
                target_server_name="target_server",
                registry_path=self.registry_path,
                task_types=["HTQC_single_point_energy"],
                report_root=self.report_root,
            )

        self.assertFalse(summary["all_passed"])
        result = summary["results"][0]
        self.assertFalse(result["success_txt_exists"])
        self.assertTrue(result["failure_txt_exists"])
        self.assertIn("without success.txt", result["failure_excerpt"])

    @override_settings(MEDIA_ROOT="/tmp")
    def test_run_fixture_replays_overrides_md_system_time_to_five_ns(self):

        registry = {
            "MDCoumpute": {
                "task_type": "MDCoumpute",
                "capability": "md_gromacs_gaussian",
                "source_server_name": "source_server",
                "source_task_id": "fixture_md_task_id",
                "source_folder_path": "/tmp/source_md_fixture",
                "source_remote_dir": "/srv/source/AutoCompute/MDCompute/Downloads/source_md_fixture",
                "task_dir_name": "source_md_fixture",
                "notebooks_to_run": list(REMOTE_REPLAY_SPECS["MDCoumpute"].notebooks_to_run),
                "func_path": REMOTE_REPLAY_SPECS["MDCoumpute"].func_path,
                "remote_target_subpath": REMOTE_REPLAY_SPECS["MDCoumpute"].remote_target_subpath,
            }
        }
        save_fixture_registry(registry, registry_path=self.registry_path)

        source_server = {
            "server_name": "source_server",
            "IP": "user@<PRIVATE_IP>:",
            "remote_target_dir": "/srv/source",
            "ssh_port": 22,
            "enabled": True,
            "capabilities": ["md_gromacs_gaussian"],
            "remote_login": "user@<PRIVATE_IP>",
            "task_limit": 3,
            "order": 0,
        }
        target_server = {
            "server_name": "target_server",
            "IP": "user@<PRIVATE_IP>:",
            "remote_target_dir": "/srv/target",
            "ssh_port": 22,
            "enabled": True,
            "capabilities": ["md_gromacs_gaussian"],
            "remote_login": "user@<PRIVATE_IP>",
            "task_limit": 3,
            "order": 1,
        }

        def fake_pull(_remote_ip, _remote_work_dir, local_dir):
            os.makedirs(local_dir, exist_ok=True)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet["A1"] = "Name"
            worksheet["B1"] = "time (ns)"
            worksheet["A2"] = "SampleA"
            worksheet["B2"] = 20
            worksheet["A3"] = "SampleB"
            worksheet["B3"] = 10
            workbook.save(os.path.join(local_dir, "System.xlsx"))
            Path(local_dir, "8_MD_process.ipynb").write_text("{}", encoding="utf-8")

        def fake_execute(*, task_func, source_dir, download_dir, task, remote_target, remote_login):  
            workbook = load_workbook(os.path.join(source_dir, "System.xlsx"))
            worksheet = workbook.active
            self.assertEqual(worksheet["B2"].value, 5)
            self.assertEqual(worksheet["B3"].value, 5)

            os.makedirs(download_dir, exist_ok=True)
            Path(download_dir, "success.txt").write_text("md replay success\n", encoding="utf-8")
            task.status = "success"
            task.save(update_fields=["status"])
            return task

        def fake_get_server(server_name, server_info_file_path=None):  
            if server_name == "source_server":
                return source_server
            if server_name == "target_server":
                return target_server
            return None

        with override_settings(MEDIA_ROOT=self.media_root), mock.patch(
            "autocompute.remote_replay_debug._get_remote_server_by_name",
            side_effect=fake_get_server,
        ), mock.patch(
            "autocompute.remote_replay_debug._check_remote_server_runtime_health",
            return_value=(True, ""),
        ), mock.patch(
            "autocompute.remote_replay_debug._pull_remote_to_local",
            side_effect=fake_pull,
        ), mock.patch(
            "autocompute.remote_replay_debug._execute_remote_task_with_result_handling",
            side_effect=fake_execute,
        ):
            summary = run_fixture_replays(
                target_server_name="target_server",
                registry_path=self.registry_path,
                task_types=["MDCoumpute"],
                report_root=self.report_root,
            )

        self.assertTrue(summary["all_passed"])
        self.assertEqual(summary["passed_count"], 1)

    def test_patched_replay_remote_execution_swaps_ssh_run_temporarily(self):

        import autocompute.remote_utils as autocompute_remote_utils_module
        import polymer.remote_utils as polymer_remote_utils_module

        original_autocompute = autocompute_remote_utils_module._ssh_run
        original_polymer = polymer_remote_utils_module._ssh_run

        with replay_module._patched_replay_remote_execution():
            self.assertIs(
                autocompute_remote_utils_module._ssh_run,
                replay_module._ssh_run_noninteractive_for_replay,
            )
            self.assertIs(
                polymer_remote_utils_module._ssh_run,
                replay_module._ssh_run_noninteractive_for_replay,
            )

        self.assertIs(autocompute_remote_utils_module._ssh_run, original_autocompute)
        self.assertIs(polymer_remote_utils_module._ssh_run, original_polymer)
